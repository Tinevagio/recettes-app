"""
Import des recettes Bristol.xlsx vers Google Sheets
----------------------------------------------------
Prérequis :
    pip install gspread google-auth openpyxl

Usage :
    python import_to_sheets.py
"""

import gspread
from google.oauth2.service_account import Credentials
import json
import openpyxl

# ─── CONFIG ───────────────────────────────────────────────
SHEET_ID   = '14AgTeaLDsXNA9NB2Si4HvRr2LtJT4pepZljAp8gVYbQ'
CREDS_FILE = 'recettes-app-490217-105d93ab1f10.json'  # à placer dans le même dossier
EXCEL_FILE = 'Bristol.xlsx'                            # à placer dans le même dossier
# ──────────────────────────────────────────────────────────

SKIP_SHEETS = {'Idées', 'LISTE', 'ASIE'}

def extract_recipes(path):
    wb = openpyxl.load_workbook(path)
    recipes = []
    recipe_id = 1

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        current_recipe = None

        for row in rows:
            if row[0] and not row[1] and not row[2]:
                if current_recipe and current_recipe['ingredients']:
                    recipes.append(current_recipe)
                    recipe_id += 1
                current_recipe = {
                    'id': recipe_id,
                    'titre': str(row[0]).strip(),
                    'feuille': sheet_name,
                    'saison': 'toute saison',
                    'source_type': 'excel',
                    'source_ref': f'Bristol.xlsx > {sheet_name}',
                    'portions': '',
                    'ingredients': []
                }
            elif current_recipe and not row[0] and (row[1] or row[2]):
                quantite   = str(row[1]).strip() if row[1] else ''
                ingredient = str(row[2]).strip() if row[2] else ''
                if ingredient and ingredient != 'None':
                    current_recipe['ingredients'].append({
                        'nom':      ingredient,
                        'quantite': quantite if quantite != 'None' else '',
                        'unite':    ''
                    })

        if current_recipe and current_recipe['ingredients']:
            recipes.append(current_recipe)
            recipe_id += 1

    return recipes


def push_to_sheets(recipes):
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=scopes)
    gc    = gspread.authorize(creds)
    sh    = gc.open_by_key(SHEET_ID)

    # ── Onglet recettes ──────────────────────────────────
    try:
        ws_r = sh.worksheet('recettes')
        ws_r.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws_r = sh.add_worksheet(title='recettes', rows=500, cols=10)

    headers_r = ['id', 'titre', 'feuille', 'saison', 'source_type', 'source_ref', 'portions']
    rows_r    = [headers_r]
    for r in recipes:
        rows_r.append([
            r['id'], r['titre'], r['feuille'], r['saison'],
            r['source_type'], r['source_ref'], r['portions']
        ])
    ws_r.update(rows_r, value_input_option='RAW')
    print(f"✅ Onglet 'recettes' : {len(recipes)} recettes importées")

    # ── Onglet ingredients ───────────────────────────────
    try:
        ws_i = sh.worksheet('ingredients')
        ws_i.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws_i = sh.add_worksheet(title='ingredients', rows=2000, cols=5)

    headers_i = ['recette_id', 'nom', 'quantite', 'unite']
    rows_i    = [headers_i]
    for r in recipes:
        for ing in r['ingredients']:
            rows_i.append([r['id'], ing['nom'], ing['quantite'], ing['unite']])
    ws_i.update(rows_i, value_input_option='RAW')
    total_ing = len(rows_i) - 1
    print(f"✅ Onglet 'ingredients' : {total_ing} ingrédients importés")


if __name__ == '__main__':
    print("📖 Lecture du fichier Excel...")
    recipes = extract_recipes(EXCEL_FILE)
    print(f"   {len(recipes)} recettes trouvées")

    print("\n📤 Import vers Google Sheets...")
    push_to_sheets(recipes)

    print("\n🎉 Import terminé !")
